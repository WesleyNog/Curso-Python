from pathlib import Path
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

workbook = Workbook()
worksheets: Worksheet = workbook.active

worksheets.cell(1, 1, 'Nome')
worksheets.cell(1, 2, 'Idade')
worksheets.cell(1, 3, 'Nota')

students = [
    ['Wesley', 14, 15.5],
    ['Naiara', 13, 9.7],
    ['Helloá', 15, 8.8],
    ['Eurides', 16, 10],
]


# for i, student_row in enumerate(students, start=2):
#     for j, student_column in enumerate(student_row, start=1):
#         worksheets.cell(i, j, student_column)

for student in students:
    worksheets.append(student)


workbook.save(WORKBOOK_PATH)